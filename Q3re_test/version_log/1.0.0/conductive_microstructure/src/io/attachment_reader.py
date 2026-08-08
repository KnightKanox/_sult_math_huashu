# 数据读取：附件.xlsx 三个分表（组1/组2/组3），每行一根介质A的两端点坐标
import openpyxl
import numpy as np


def read_attachment(path):
    """读取附件Excel，返回 {分表名: (N,6) 数组}，每行 [p_x,p_y,p_z,q_x,q_y,q_z]。"""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = []
        for r in ws.iter_rows(values_only=True):
            # 跳过表头（首列为非数值文本）
            if r[0] is None or isinstance(r[0], str):
                continue
            rows.append([float(v) for v in r[:6]])
        result[sheet_name] = np.array(rows, dtype=float)
    wb.close()
    return result
